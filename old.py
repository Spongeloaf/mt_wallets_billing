from mailmerge import MailMerge
import sys
import smtplib
import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook




def prepare_bill(tenant_num):
    """ Generates the bills """
    document = MailMerge(billingTemplate)  # rename this so it's easier to use later. Possibly inherits some shit form MailMerge?
    tenant = Tenant()
    
    ###################      CONTINUE WORK HERE!!!!    #################
    # You were fixing all these properties into an object. You also need a global parameters object with billing strings and sheets and such.

    tenant.email = wb.worksheets[1].cell(column=3, row=tenant_num).value  # assign a tenant email_addr
    tenant.name = format_values(wb[sheets[1]].cell(column=2, row=tenant_num).value)
    tenant.roomsheet = wb.worksheets[1].cell(column=1, row=tenant_num).value  # assign a tenant room number based on each tenants row in the sheet.
    tenant.date = format_values(billingDate.strftime('%d %B %Y'))
    tenant.billing = billingDate.strftime('%B %Y')
    tenant.other = roomSheet.cell(column=11, row=billingCycle).value
    tenant.internet = roomSheet.cell(column=7, row=billingCycle).value
    tenant.room = roomSheet.cell(column=5, row=billingCycle).value
    tenant.total = roomSheet.cell(column=2, row=billingCycle).value
    tenant.electricity = roomSheet.cell(column=8, row=billingCycle).value
    tenant.gas = roomSheet.cell(column=6, row=billingCycle).value


    document.merge(
        #move all value resolution functions the fuck out of here and build a new matrix for them. Add tenant names, dates etc.
        name=format_values(mergeName),
        date=format_values(billingDate.strftime('%d %B %Y')),
        billing=format_values(mergeBilling),
        other=format_values(mergeOther),
        internet=format_values(mergeInternet),
        room=format_values(mergeRoom),
        total=format_values(mergeTotal),
        electricity=format_values(mergeElectricity),
        gas=format_values(mergeGas)
        )

    document.write("{}s Rent for {}.docx".format(format_values(mergeName), mergeBilling)) # Save the new bill




# init
billingTemplate = 'billingTemplate.docx'                        # This shit here is the template for generated bills
billingDatabase = 'billingv2.xlsx'                              # This shit here is the database for billing
wb = load_workbook(filename=billingDatabase, guess_types=True, data_only=True)  # load the workbook, and label the workbook container as "wb"
sheets = wb.sheetnames      # fetch the list of worksheet names and store them here
billingWorksheet = wb[sheets[0]]  # set the worksheet we are searching. In this case, the first one.
tennant_sheet = wb[sheets[1]]
roomCount = wb.worksheets[1].max_row  # figure out how many tenants there are by counting rows in the tenant sheet
roomCount = roomCount + 1  # offset by 1 to account for starting at 0

# user input
billingString = input("Enter billing string: ")
runMode = input("Please Choose an option:\n1: Generate and Send\n2: Generate Only\n3: Send Only")


if runMode == "1":
    print('Generate and Send')
    print("billing string:{0}".format(billingString))
    print("-------")
    for billingCycle in range(1, 900):          # We will search the first column of the first sheet for the current date string
        if billingWorksheet.cell(column=1, row=billingCycle).value == billingString:    # check for match in each row.
            print ("Preparing billing cycle {}.....".format(billingCycle))                             # print billing cycle string for debugging
            for tennant in range(1, tennant_sheet.max_row):
                prepare_bill(tennant)
                send_bill(tennant)

if runMode == "2":
    print('Generate Only')
    print("billing string:{0}".format(billingString))
    print("-------")
    for billingCycle in range(1, 900):          # We will search the first column of the first sheet for the current date string
        if billingWorksheet.cell(column=1, row=billingCycle).value == billingString:    # check for match in each row.
            print ("Preparing billing cycle {}.....".format(billingCycle))                             # print billing cycle string for debugging
            for tennant in range(1, tennant_sheet.max_row):
                prepare_bill(tennant)

if runMode == "3":
    print('Send Only')
    print("billing string:{0}".format(billingString))
    print("-------")
    for billingCycle in range(1, 900):          # We will search the first column of the first sheet for the current date string
        if billingWorksheet.cell(column=1, row=billingCycle).value == billingString:    # check for match in each row.
            print ("Preparing billing cycle {}.....".format(billingCycle))                             # print billing cycle string for debugging
            for tennant in range(1, tennant_sheet.max_row):
                send_bill(tennant)

sys.exit("Done")